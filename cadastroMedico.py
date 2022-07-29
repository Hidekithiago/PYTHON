#   82@43Bqa


#pip3 install xlrd  install imap-tools install webdriver-manager install selenium install mysql-connector install pandas install chromedriver-autoinstaller install webdrivermanager install openpyxl install opencv-python install pyexcel install pyexcel-xls install pyexcel-xlsx
import os,sys
#import insereMedicoBd as bd
from audioop import cross
from cmath import nan
from logging import exception
from pydoc import cram
from cv2 import line
# from socket import J1939_PGN_ADDRESS_CLAIMED
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
import requests
import time
import mysql.connector
import json
from datetime import date, datetime
import shutil
from zipfile import ZipFile
import pandas as pd

import smtplib
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from pathlib import Path
import openpyxl
import pyexcel as p
#Conexão com o banco de dados
# conn = mysql.connector.connect(host="frotaleve.cqth4ctrsht2.us-east-1.rds.amazonaws.com",
#                               user="sys.atestadosdev", password="Js*23676@3", db="atestadosdev2")  # Ambiente de Homologacao

dir_download = r"/home/quaestum/cadastroMedicoSOC/download" # Pega o diretorio da pasta de Download Linux
#dir_download = str(Path.home() / "Downloads") # Pega o diretorio da pasta de Download Windows
download =  executable_path= dir_download+r'//Consulta.zip' #Diretorio de Download da API
#download =  executable_path= dir_download+r'\Consulta.zip' #Diretorio de Download da API
newAdress = executable_path=r'/home/quaestum/cadastroMedicoSOC/downloadAPI' #Diretorio onde a API será alocada depois do Download
#newAdress = executable_path=r'C:\quaestum\python\cadastroMedico\downloadAPI' #Diretorio onde a API será alocada depois do Download
extract = executable_path=newAdress+r'//Consulta.zip' #Arquivo que será descompactado
#ChromeOptions
chromeOptions = webdriver.ChromeOptions()
chromeOptions.add_argument("--disable-notifications")
chromeOptions.add_argument('--no-sandbox')
chromeOptions.add_argument('--verbose')
chromeOptions.add_experimental_option("prefs", {
        "download.default_directory": dir_download,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing_for_trusted_sources_enabled": False,
        "safebrowsing.enabled": False
    })

chromeOptions.add_argument('--disable-gpu')
chromeOptions.add_argument('--disable-software-rasterizer')
#chromeOptions.add_argument('--headless')     
#driver = webdriver.Chrome(executable_path=r"/home/quaestum/cadastroMedicoSOC/chromedrive/chromedriver", options=chromeOptions) #Iniciando o driver do Chrome
driver = webdriver.Chrome((ChromeDriverManager().install()), options=chromeOptions)


cro_crm = ''
siglaConselho = ''
ufMed = ''
codPessoa = ''
nomeMedico = ''
codPessoaBD = ''

def setId(b1):
    #Identifica qual e o valor de cada botao no login
    bt_1 = driver.find_element(By.ID,"bt_1").get_attribute("value")
    bt_2 = driver.find_element(By.ID,"bt_2").get_attribute("value")
    bt_3 = driver.find_element(By.ID,"bt_3").get_attribute("value")
    bt_4 = driver.find_element(By.ID,"bt_4").get_attribute("value")
    bt_5 = driver.find_element(By.ID,"bt_5").get_attribute("value")
    bt_6 = driver.find_element(By.ID,"bt_6").get_attribute("value")
    bt_7 = driver.find_element(By.ID,"bt_7").get_attribute("value")
    bt_8 = driver.find_element(By.ID,"bt_8").get_attribute("value")
    bt_9 = driver.find_element(By.ID,"bt_9").get_attribute("value")
    bt_0 = driver.find_element(By.ID,"bt_0").get_attribute("value")
        #Atribui os valores do ID nas variaveis
    if bt_1 == b1: return 'bt_1'
    elif bt_2 == b1: return 'bt_2'
    elif bt_3 == b1: return 'bt_3'
    elif bt_4 == b1: return 'bt_4'
    elif bt_5 == b1: return 'bt_5'
    elif bt_6 == b1: return 'bt_6'
    elif bt_7 == b1: return 'bt_7'
    elif bt_8 == b1: return 'bt_8'
    elif bt_9 == b1: return 'bt_9'
    elif bt_0 == b1: return 'bt_0'

#Realiza o Login no sistema do SOC
def LoginSoc():
    print('\n\n\n\n\nIniciando o SOC')
    driver.get('https://sistema.soc.com.br/WebSoc/')
    time.sleep(2)
    print('Acessou o SOC')
    driver.maximize_window()
    time.sleep(2)

    #Insere as credenciais

    driver.get("https://sistema.soc.com.br/WebSoc/")#Vai para o endereco do SOC    
    #print( driver.page_source)
    usuario = driver.find_element(By.XPATH, "/html/body/div[1]/div/form/div[2]/div[1]/div[1]/input")
    usuario.send_keys("thideki")
    time.sleep(2)
    senha = driver.find_element(By.XPATH,"/html/body/div[1]/div/form/div[2]/div[1]/div[2]/input")
    senha.send_keys("abc123**57951221")

    #Seleciona os digitos do ID
    driver.find_element(By.ID,setId('6')).click()
    driver.find_element(By.ID,setId('6')).click()
    driver.find_element(By.ID,setId('3')).click()
    driver.find_element(By.ID,setId('4')).click()
    driver.find_element(By.ID,"bt_entrar").click()

#Realiza o download da API
def downloadAPI():
    #Caso apareca alguma mensagem do SOC
    try:
        driver.find_element(By.XPATH,"/html/body/div[11]/div[2]/p/input").click()
        driver.find_element(By.XPATH,"/html/body/div[11]/div[2]/p/span/a").click()
    except:pass
    #Acessa a pagina para gerar o relatorio
    driver.find_element(By.XPATH,"/html/body/div[1]/a/img").click()
    driver.find_element(By.XPATH,"/html/body/div[7]/div[4]").click()
    driver.find_element(By.XPATH,"/html/body/div[7]/div[4]/div/table/tbody/tr[1]/td/div/table/tbody/tr").click()
    driver.find_element(By.XPATH,"/html/body/div[7]/div[4]/div/table/tbody/tr[1]/td/div/table/tbody/tr/td[3]/div/table/tbody/tr/td/span").click()
    
    #Seleciona as opções do relatorio(CheckBox)
    driver.switch_to.frame(1)
    driver.find_element(By.ID,'usuario').click()
    driver.find_element(By.ID,'examinador').click()
    driver.find_element(By.ID,'solicitante').click()
    driver.find_element(By.ID,'gerente').click()
    driver.find_element(By.ID,'profissional').click()
    driver.find_element(By.ID,'emissorAso').click()
    driver.find_element(By.ID,'responsavel').click()
    driver.find_element(By.ID,'palestrante').click()
    print('Fez login')
    time.sleep(1)
    #Limpa o e-mail para envio do relatorio
    driver.find_element(By.XPATH,"/html/body/div[2]/div/form[1]/fieldset[2]/p/img[3]").click()
    #digita o email para enviar email
    driver.find_element(By.XPATH,"/html/body/div[2]/div/form[1]/fieldset[2]/p/input").send_keys("processosoc12@gmail.com")    
    time.sleep(3)
    #Gera relatorio
    driver.find_element(By.XPATH,"/html/body/div[2]/div/form[1]/age_nao_gravar/div[1]/table/tbody/tr/td[3]/a/img").click()
    time.sleep(1)
    # driver.find_element(By.XPATH,"/html/body/div[2]/div[3]/a").click()
    # driver.find_element(By.XPATH,"/html/body/div[5]/table/tbody/tr[2]/td[3]/a/img").click()
    # driver.find_element(By.XPATH,"/html/body/div[5]/table/tbody/tr[2]/td[4]/a").click()

    #cards = driver.find_elements_by_class_name("pedProcMsg")


    #Captura o código da consulta
    cdConsulta = driver.find_element(By.XPATH,"/html/body/div[2]/div/form[1]/span/div/p[1]/b").text
    print(cdConsulta)
    driver.find_element(By.XPATH,"/html/body/div[2]/div/form[1]/span/div/p[3]/a").click()
    nuPedido = driver.find_element(By.XPATH,"/html/body/div[2]/div/form[1]/table/tbody/tr[1]/td/table/tbody/tr[5]/td[3]/input")
    nuPedido.send_keys(cdConsulta)
    driver.find_element(By.XPATH,"/html/body/div[2]/div/form[1]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[3]/a/img").click()
    print('Aguardando aparecer o download')
    time.sleep(90)
    driver.find_element(By.XPATH,"/html/body/div[2]/div/form[1]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[3]/a/img").click()
        
    driver.find_element(By.XPATH,"/html/body/div[2]/div/form[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td[10]/div[2]/a").click()
    time.sleep(3)
    driver.find_element(By.XPATH,"/html/body/div[2]/div/form[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td[10]/div[2]/a").click()
    time.sleep(3)
    driver.find_element(By.XPATH,"/html/body/div[2]/div/form[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td[10]/div[2]/a").click()
    print("Download ok")
    time.sleep(15)
    driver.close()

#Extração .zip e move para a pasta de leitura NewAdresss
def formataAPI():
    shutil.move(download, newAdress)
    print("Move ok")
    z = ZipFile(extract, 'r')
    z.extractall(newAdress)
    z.close()
    print(" ok")

#Leitura da API e lógica de inserção no banco de dados
def lerAPI():

    arq = os.listdir(newAdress)
    newAdress.format(".xls").endswith(".xls")

    file_name = executable_path= newAdress + r'//'+str(arq[0])
    #file_name = executable_path= newAdress + r'\\'+str(arq[0])

    i = ''
    resultPlan = pd.read_excel(file_name,
    sheet_name="Exportar Modelo Pessoa", usecols=[0,2,17,18,19])
    row = resultPlan.shape[0]
    for i in range (row):
        codPessoa = resultPlan.values [i][0]
        nomeMedico = resultPlan.values [i][1]
        siglaConselho = resultPlan.values [i][2]
        cro_crm = str(resultPlan.values[i][3])
        ufMed = str(resultPlan.values [i][4])

        print(codPessoa)
        print(cro_crm)
        print(nomeMedico)
        print(siglaConselho)
        print(ufMed)

        if ufMed == 'nan':
            ufMed = ""

        if "_" in cro_crm:
            siglaConselho = 'erro'

        try:
            if siglaConselho == 'cro' or siglaConselho == 'CRO':
                query = "SELECT * FROM medicoCRO where cd_pessoa = '"+codPessoa+"'"
                for result in cursor.execute(query, multi=True):
                    rowBD = result.fetchall()
                    if len(rowBD) > 0:
                        print("já existe")
                    else:
                        query3 = "INSERT INTO medicoCRO(nome, cro, uf, cd_pessoa)VALUES('"+str(nomeMedico)+"','"+str(cro_crm)+"','"+str(ufMed)+"','"+str(codPessoa)+"')"
                        cursor.execute(query3)
                        conn.commit()
            else:
                print("não é cro")
        except:
            print("erro ao inserir cro ou consultar")
            time.sleep(20)
            
        try:
            if siglaConselho == 'crm' or siglaConselho == 'CRM':
                query2 = "SELECT * FROM medicoCRM where cd_pessoa = '"+codPessoa+"'"
                for result in cursor.execute(query2, multi=True):
                    rowBD = result.fetchall()
                    if len(rowBD) > 0:
                        print("já existe")
                    else:
                        query4 = "INSERT INTO medicoCRM(nome, crm, uf, cd_pessoa)VALUES('"+str(nomeMedico)+"','"+str(cro_crm)+"','"+str(ufMed)+"','"+str(codPessoa)+"')"
                        cursor.execute(query4)
                        conn.commit()
            else:
                print("não é crm")
        except exception as a:

            print("erro ao inserir crm ou consultar")
            print(a)
            time.sleep(20)

def createQuery():    
    arq = os.listdir(newAdress)
    newAdress.format(".xls").endswith(".xls")
    
    print(arq[1])
    if arq[1] == 'Consulta.zip':    
        file_name = r'/home/quaestum/cadastroMedicoSOC/downloadAPI//'+str(arq[0])
    else:
        file_name = r'/home/quaestum/cadastroMedicoSOC/downloadAPI//'+str(arq[1])
    #file_name = r'C:\quaestum\python\cadastroMedico\downloadAPI\\'+str(arq[0])
    
    path = Path(file_name+str('x'))
    if path.is_file():
        file_name = file_name+str('x')
    else:        
        p.save_book_as(file_name=file_name,dest_file_name=file_name+str('x'))
        file_name = file_name+str('x')
    dirqueryCrm = r'/home/quaestum/cadastroMedicoSOC/querycrm.txt'
    dirqueryCro = r'/home/quaestum/cadastroMedicoSOC/querycro.txt'
    #dirqueryCrm =r'C:\quaestum\python\cadastroMedico\querycrm.txt'
    #dirqueryCro = r'C:\quaestum\python\cadastroMedico\querycro.txt'
    queryCrm = readQuery(dirqueryCrm, tableCRM, 'crm')
    queryCro = readQuery(dirqueryCro, tableCRO, 'cro')
    i = ''
    
    leituraXls('Exportar Modelo Pessoa', file_name, dirqueryCro, dirqueryCrm)
    leituraXls('Exportar Modelo Pessoa1', file_name, dirqueryCro, dirqueryCrm)

    line = ''
    with open(dirqueryCrm) as f:
        lines = f.readlines()
        for x in lines:
            line += x
    line = line.replace('(nome, crm, uf, cd_pessoa)VALUES ,','(nome, crm, uf, cd_pessoa)VALUES ')
    f = open(dirqueryCrm, "w")
    f.write(line)
    f.close()
    line = ''
    with open(dirqueryCro) as f:
        lines = f.readlines()
        for x in lines:
            line += x
    line = line.replace('(nome, cro, uf, cd_pessoa)VALUES ,','(nome, cro, uf, cd_pessoa)VALUES ')
    f = open(dirqueryCro, "w")
    f.write(line)
    f.close()

def leituraXls(worksheet, file_name, dirqueryCro, dirqueryCrm):
    book = openpyxl.load_workbook(file_name)    
    sheet_obj = book[worksheet]
    for x in range (sheet_obj.max_row):
        print(x)
        codPessoa = sheet_obj.cell(row = (x+1), column = 1).value
        nomeMedico = sheet_obj.cell(row = (x+1), column = 3).value
        siglaConselho = sheet_obj.cell(row = (x+1), column = 18).value
        cro_crm = sheet_obj.cell(row = (x+1), column = 19).value
        ufMed = sheet_obj.cell(row = (x+1), column = 20).value

        print(codPessoa)
        print(cro_crm)
        print(nomeMedico)
        print(siglaConselho)
        print(ufMed)
        
        if codPessoa is None: codPessoa = ''
        if nomeMedico is None: nomeMedico = ''
        if siglaConselho is None: siglaConselho = ''
        if cro_crm is None: cro_crm = ''
        if ufMed is None: ufMed = ''

        
        if ufMed == 'nan':
            ufMed = ""
        if "_" in cro_crm:
            siglaConselho = 'erro'
        try:
            if siglaConselho == 'cro' or siglaConselho == 'CRO':
                f = open(dirqueryCro, "a")
                f.write(", ('"+str(nomeMedico.replace("'",""))+"','"+str(cro_crm.replace("'",""))+"','"+str(ufMed.replace("'",""))+"','"+str(codPessoa.replace("'",""))+"')")
                f.close()
            elif siglaConselho == 'crm' or siglaConselho == 'CRM':
                f = open(dirqueryCrm, "a")
                f.write(", ('"+str(nomeMedico.replace("'",""))+"','"+str(cro_crm.replace("'",""))+"','"+str(ufMed.replace("'",""))+"','"+str(codPessoa.replace("'",""))+"')")
                f.close()
            #sheet_obj.delete_rows(x)
        except Exception as e:
            print (e)
        #book.save(file_name)
    
def readQuery(dir, type, crm_cro):
    try:
        with open(dir) as f:
            lines = f.readlines()
            for x in lines:
                line += x
        return line
    except:
        f = open(dir, "w")
        f.write("INSERT INTO %s(nome, %s, uf, cd_pessoa)VALUES " %(type, crm_cro))
        f.close()
        f = open(dir, "r")
        return(f.read())

def removeFiles(path):    
    dir = os.listdir(path)
    for file in dir:
        os.remove(path+'//'+file)
        #os.remove(path+'\\'+file)

def getAttEmail():
    # pegar emails de um remetente para um destinatário
    gmail_user = "processosoc12@gmail.com"
    gmail_password = "ofbxufqjbhavqmkm"

    try:
        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.ehlo()
        #server.starttls()
        server.login(gmail_user, gmail_password)
        server.sendmail(gmail_user, gmail_user, 'email_text')
        server.close()

        print ('Email sent!')
    except Exception as a:
        print(a)
        print ('Something went wrong...')
        
tableCRM = 'medicoCRM'
tableCRO = 'medicoCRO'
#removeFiles(newAdress)
#removeFiles(r'C:\quaestum\python\cadastroMedico\downloadAPI')
LoginSoc ()
downloadAPI()
formataAPI()
createQuery()
command = 'python3 "/home/quaestum/cadastroMedicoSOC/insereMedicoBd.py"'
#command = 'py "C:\\quaestum\\insereMedicoBd.py"'
if os.system(command) == 0: print("Executado com sucesso.")
else:print("Erro ao executar o comando.")
#bd.run('C:\quaestum\python\cadastroMedico\querycrm.txt', tableCRM)

for f in os.listdir(newAdress):
    os.remove(os.path.join(newAdress, f))

print("Sucess")