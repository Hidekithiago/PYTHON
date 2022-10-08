'''| Comando para instalar as bibliotecas do projeto |'''
#pip3 install selenium
import os,sys
'''| Biblioteca para ler a planilha |'''
import openpyxl
'''| Biblioteca para pegar o diretorio das pastas |'''
from pathlib import Path
'''| Biblioteca para adicionar um delay na execucao |'''
import time
'''| Biblioteca pegar a data e hora atual |'''
from datetime import datetime




def readSheet():
    path = ''
    wb_obj = openpyxl.load_workbook(path) 


    print(wb_obj.get_sheet_names())
    # Read the active sheet:
    sheet = wb_obj.active
    print(sheet)
    #print(sheet["C2"].value)
    print(sheet.max_row, sheet.max_column)
    col_names = []
    for column in sheet.iter_cols(1, sheet.max_column):
        col_names.append(column[0].value)




    print(col_names)
    data = {}


    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            data[row[1]] = []
            data[row[2]] = []
            data[row[3]] = []


        else:
            data['a'].append(row[1])
            data['b'].append(row[2])
            data['c'].append(row[3])
    print(data['a'])


if __name__ == "__main__": #Inicio do programa
    print("Inicio do programa")
    inicio = datetime.now()


    readSheet()


    fim = datetime.now()
    print('\n\nTempo de execucao do programa: %s'%(fim-inicio))

