#pip install openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook


def leituraExcel():
    arquivo_excel = load_workbook(dirExcelRead)
    planilha1 = arquivo_excel.active
    max_linha = planilha1.max_row
    max_coluna = planilha1.max_column
    for i in range(1, max_linha + 1):
        for j in range(1, max_coluna + 1):
            print(planilha1.cell(row=i, column=j).value, end=" - ")
        print('\n')
            
if __name__ == "__main__":
    dirExcelRead = r'C:\quaestum\robo\MODELO_PADRAO_EQUIPES (2).xlsx'
    leituraExcel()
